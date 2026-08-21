"""
The single source of truth for the student bulk-import file format.

The importer, the downloadable Excel template and the on-screen column guide
are all driven from this list, so what an admin is told to type is exactly
what the parser accepts.
"""

# key         — the column heading (headings are matched case-insensitively,
#               and spaces/hyphens/slashes are treated as underscores)
# label       — human heading used in the template
# required    — must be filled in on every row
# rule        — how to fill the cell in
# example     — a filled-in value
COLUMNS = [
    {
        "key": "reg_no", "label": "reg_no", "required": True,
        "rule": "Registration number. Must be unique — a number already in the system, or repeated twice in this file, is rejected.",
        "example": "AN-2026-001",
    },
    {
        "key": "first_name", "label": "first_name", "required": True,
        "rule": "Student's first name.", "example": "Asha",
    },
    {
        "key": "last_name", "label": "last_name", "required": True,
        "rule": "Student's surname.", "example": "Said",
    },
    {
        "key": "middle_name", "label": "middle_name", "required": False,
        "rule": "Middle name. Leave blank if none.", "example": "Juma",
    },
    {
        "key": "date_of_birth", "label": "date_of_birth", "required": True,
        "rule": "Date of birth. A real Excel date cell works, or type YYYY-MM-DD / DD-MM-YYYY / DD.MM.YYYY.",
        "example": "2014-03-02",
    },
    {
        "key": "gender", "label": "gender", "required": True,
        "rule": "Male or Female (M / F also accepted; capitalisation does not matter).",
        "example": "Female",
    },
    {
        "key": "admission_date", "label": "admission_date", "required": True,
        "rule": "Date the student joined the school. Same date formats as date_of_birth.",
        "example": "2026-01-13",
    },
    {
        "key": "academic_session", "label": "academic_session", "required": False,
        "rule": "Academic year. Defaults to 2026 when blank.", "example": "2026",
    },
    {
        "key": "student_class", "label": "student_class", "required": False,
        "rule": "Class name exactly as it appears under Academics → Classes (e.g. Form 1A). The class must already exist. Leave blank to place the student later.",
        "example": "Form 1A",
    },
    {
        "key": "student_type", "label": "student_type", "required": False,
        "rule": "Day or Boarding. Defaults to Day when blank.", "example": "Day",
    },
    {
        "key": "status", "label": "status", "required": False,
        "rule": "active, suspended, graduated or withdrawn. Defaults to active.",
        "example": "active",
    },
    {
        "key": "fee_status", "label": "fee_status", "required": False,
        "rule": "paid, partial or unpaid. Defaults to unpaid.", "example": "unpaid",
    },
    {
        "key": "blood_group", "label": "blood_group", "required": False,
        "rule": "Blood group, e.g. O+.", "example": "O+",
    },
    {
        "key": "religion", "label": "religion", "required": False,
        "rule": "Religion.", "example": "Islam",
    },
    {
        "key": "state_of_origin", "label": "state_of_origin", "required": False,
        "rule": "Region or state of origin.", "example": "Zanzibar Urban",
    },
    {
        "key": "residential_address", "label": "residential_address", "required": False,
        "rule": "Home address.", "example": "Kisauni, West B District",
    },
    {
        "key": "previous_school", "label": "previous_school", "required": False,
        "rule": "Name of the school the student attended before.", "example": "Mwanakwerekwe Primary",
    },
    {
        "key": "previous_class", "label": "previous_class", "required": False,
        "rule": "Class the student was in at that school.", "example": "Std 7",
    },
    {
        "key": "is_orphan", "label": "is_orphan", "required": False,
        "rule": "TRUE or FALSE (yes/no/1/0 also accepted). Blank means FALSE.",
        "example": "FALSE",
    },
    {
        "key": "has_disability", "label": "has_disability", "required": False,
        "rule": "TRUE or FALSE. If TRUE, disability_details must be filled in.",
        "example": "FALSE",
    },
    {
        "key": "disability_details", "label": "disability_details", "required": False,
        "rule": "Required only when has_disability is TRUE. Describe the disability and the support needed.",
        "example": "",
    },
    {
        "key": "donor", "label": "donor", "required": False,
        "rule": "Sponsor name exactly as it appears under Donors & Sponsors. The donor must already exist.",
        "example": "",
    },
    {
        "key": "donor_number", "label": "donor_number", "required": False,
        "rule": "Sponsor's reference number for this student.", "example": "",
    },
    {
        "key": "parent_name", "label": "parent_name", "required": False,
        "rule": "Parent or guardian's full name. If you fill in any parent column you must fill in parent_name, parent_phone AND relationship.",
        "example": "Fatma Said",
    },
    {
        "key": "parent_phone", "label": "parent_phone", "required": False,
        "rule": "Parent's phone number. Required whenever parent_name is given.",
        "example": "+255 774 221 707",
    },
    {
        "key": "relationship", "label": "relationship", "required": False,
        "rule": "How the guardian is related to the student, e.g. Mother. Required whenever parent_name is given.",
        "example": "Mother",
    },
    {
        "key": "parent_email", "label": "parent_email", "required": False,
        "rule": "Parent's e-mail. Used later to link their parent portal login.",
        "example": "fatma@example.com",
    },
    {
        "key": "occupation", "label": "occupation", "required": False,
        "rule": "Parent's occupation.", "example": "Teacher",
    },
    {
        "key": "office_address", "label": "office_address", "required": False,
        "rule": "Parent's work address.", "example": "",
    },
    {
        "key": "home_address", "label": "home_address", "required": False,
        "rule": "Parent's home address.", "example": "",
    },
    {
        "key": "guardian2_name", "label": "guardian2_name", "required": False,
        "rule": "A second guardian — the father, uncle or elder brother alongside the mother. Fill in guardian2_name, guardian2_phone AND guardian2_relationship together. They can be given their own parent-portal login afterwards.",
        "example": "Said Juma",
    },
    {
        "key": "guardian2_phone", "label": "guardian2_phone", "required": False,
        "rule": "Second guardian's phone number. Required whenever guardian2_name is given.",
        "example": "+255 777 397 422",
    },
    {
        "key": "guardian2_relationship", "label": "guardian2_relationship", "required": False,
        "rule": "How the second guardian is related to the student, e.g. Father, Uncle, Brother. Required whenever guardian2_name is given.",
        "example": "Father",
    },
    {
        "key": "guardian2_email", "label": "guardian2_email", "required": False,
        "rule": "Second guardian's e-mail. Used to link their portal login.",
        "example": "said@example.com",
    },
    {
        "key": "guardian2_occupation", "label": "guardian2_occupation", "required": False,
        "rule": "Second guardian's occupation.", "example": "",
    },
]

REQUIRED_KEYS = [c["key"] for c in COLUMNS if c["required"]]


def build_template_workbook():
    """An .xlsx with the headings, one worked example, and a column guide."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Students"
    sheet.append([c["label"] for c in COLUMNS])
    sheet.append([c["example"] for c in COLUMNS])

    header_font = Font(bold=True, color="FFFFFF")
    required_fill = PatternFill("solid", fgColor="1F6F43")
    optional_fill = PatternFill("solid", fgColor="4A5568")
    for column_index, spec in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = header_font
        cell.fill = required_fill if spec["required"] else optional_fill
        cell.alignment = Alignment(horizontal="center")
        width = max(len(spec["label"]) + 4, len(str(spec["example"])) + 4, 14)
        sheet.column_dimensions[cell.column_letter].width = min(width, 32)
    sheet.freeze_panes = "A2"

    guide = workbook.create_sheet("Column guide")
    guide.append(["Column", "Required?", "How to fill it in", "Example"])
    for spec in COLUMNS:
        guide.append([
            spec["label"],
            "Required" if spec["required"] else "Optional",
            spec["rule"],
            str(spec["example"]),
        ])
    for cell in guide[1]:
        cell.font = header_font
        cell.fill = optional_fill
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 12
    guide.column_dimensions["C"].width = 90
    guide.column_dimensions["D"].width = 26
    for row in guide.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
    guide.freeze_panes = "A2"

    return workbook
