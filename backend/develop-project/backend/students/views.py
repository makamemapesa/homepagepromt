import csv
import io
import datetime
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.dateparse import parse_date
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from core.permissions import (
    IsSuperAdmin,
    IsSuperAdminOrAdmin,
    IsSchoolStaffOrParentReadOnly,
    IsTeacherOrAdminOrParentReadOnly,
)
from core.utils import get_teacher_class_ids, get_teacher_for, get_user_role
from donors.models import Donor
from .models import Student, StudentDocument, AcademicHistory
from .serializers import (
    StudentListSerializer,
    StudentDetailSerializer,
    StudentCreateSerializer,
    StudentDocumentSerializer,
    AcademicHistorySerializer,
)


class StudentViewSet(viewsets.ModelViewSet):
    """
    Student management with role-based access:
    - Super Admin/Admin: Full access to all students
    - Teacher: View students in their classes only
    - Parent: View only their own children
    - Accountant: View all students (for fee management)
    """
    queryset = Student.objects.select_related("student_class", "donor").all()
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["status", "fee_status", "student_type", "gender", "is_orphan", "has_disability", "student_class"]
    search_fields = ["first_name", "last_name", "reg_no", "student_class__name"]
    ordering_fields = ["last_name", "first_name", "admission_date", "reg_no"]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_permissions(self):
        """
        Set permissions based on action:
        - Create/Update/Delete: Only admins
        - Import students: Only super admins
        - View: Admins, teachers (their classes), parents (their children), accountants
        """
        if self.action in ('import_students', 'import_columns', 'import_template'):
            return [IsSuperAdmin()]
        if self.action in ['create', 'destroy', 'promote']:
            return [IsSuperAdminOrAdmin()]
        elif self.action in ['update', 'partial_update']:
            return [IsSuperAdminOrAdmin()]
        # Read access: accountants included so they can look a student up when
        # recording a payment. get_queryset() below scopes what each role sees.
        return [IsSchoolStaffOrParentReadOnly()]

    def get_queryset(self):
        """
        Filter students based on user role.
        """
        user = self.request.user
        role = get_user_role(user)

        if role in ['super_admin', 'admin', 'accountant']:
            # Admins and accountants see all students
            return Student.objects.select_related("student_class", "donor").all()
        
        elif role == 'teacher':
            # Teachers see students in their classes (homeroom + subject-assigned)
            all_ids = get_teacher_class_ids(get_teacher_for(user))
            return Student.objects.filter(
                student_class_id__in=all_ids
            ).select_related("student_class", "donor")
        
        elif role == 'parent':
            # Parents see only their own children
            return Student.objects.filter(
                guardians__user=user
            ).select_related("student_class", "donor").distinct()
        
        # Default: no access
        return Student.objects.none()

    def get_serializer_class(self):
        if self.action == "list":
            return StudentListSerializer
        if self.action in ["create", "update", "partial_update"]:
            return StudentCreateSerializer
        return StudentDetailSerializer

    def _normalize_header(self, header):
        if not header:
            return ""
        return str(header).strip().lower().replace(" ", "_").replace("-", "_").replace("/", "_")

    def _parse_bool(self, value):
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        return normalized in ["1", "true", "yes", "y", "t", "on"]

    def _parse_date(self, value):
        if value is None or str(value).strip() == "":
            return None
        # Excel hands back real date cells as datetime objects; the serializer's
        # DateField rejects those, so narrow them to a plain date first.
        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value
        normalized = str(value).strip()
        parsed = parse_date(normalized)
        if parsed:
            return parsed
        for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"]:
            try:
                return datetime.datetime.strptime(normalized, fmt).date()
            except ValueError:
                continue
        raise ValidationError(f"Invalid date format: {normalized}")

    def _parse_student_class(self, value):
        if value is None or str(value).strip() == "":
            return None
        normalized = str(value).strip()
        from academics.models import Class
        if normalized.isdigit():
            try:
                return int(normalized)
            except ValueError:
                pass
        try:
            cls = Class.objects.filter(name__iexact=normalized).first()
            if cls:
                return cls.id
        except Exception:
            pass
        if normalized.isdigit():
            if Class.objects.filter(pk=int(normalized)).exists():
                return int(normalized)
        raise ValidationError(f"Student class not found: {normalized}")

    def _parse_donor(self, value):
        if value is None or str(value).strip() == "":
            return None
        normalized = str(value).strip()
        if normalized.isdigit():
            donor = Donor.objects.filter(pk=int(normalized)).first()
            if donor:
                return donor.id
        donor = Donor.objects.filter(name__iexact=normalized).first()
        if donor:
            return donor.id
        raise ValidationError(f"Donor not found: {normalized}")

    @staticmethod
    def _row_is_blank(row):
        """True when every cell in the row is empty — Excel adds these freely."""
        return all(str(value).strip() == "" for value in row.values() if value is not None) and not any(
            value not in (None, "") for value in row.values()
        )

    def _parse_import_rows(self, file):
        """Read the upload into [(spreadsheet_row_number, {column: value}), ...].

        Blank rows are dropped rather than reported as errors, because
        spreadsheets routinely carry trailing empties.
        """
        file_name = getattr(file, "name", "")
        numbered = []

        if file_name.lower().endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise ValidationError(
                    "Excel support is not installed on the server (openpyxl). "
                    "Save the file as .csv and upload that instead, or ask your "
                    "administrator to run: pip install -r requirements.txt"
                )
            workbook = load_workbook(filename=file, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [self._normalize_header(c) for c in rows[0]]
            if not any(headers):
                raise ValidationError("The first row must contain the column headings.")
            for offset, row in enumerate(rows[1:], start=2):
                parsed = {
                    headers[idx]: cell
                    for idx, cell in enumerate(row)
                    if idx < len(headers) and headers[idx]
                }
                numbered.append((offset, parsed))
        else:
            text = file.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise ValidationError("The first row must contain the column headings.")
            for offset, raw_row in enumerate(reader, start=2):
                parsed = {}
                for raw_key, value in raw_row.items():
                    key = self._normalize_header(raw_key)
                    if key:
                        parsed[key] = value
                numbered.append((offset, parsed))

        return [(number, row) for number, row in numbered if not self._row_is_blank(row)]

    def _build_import_payload(self, row):
        row = {self._normalize_header(key): value for key, value in row.items()}
        def get_val(key):
            return row.get(self._normalize_header(key), "")

        payload = {
            "reg_no": str(get_val("reg_no") or get_val("registration_number") or "").strip(),
            "first_name": str(get_val("first_name") or "").strip(),
            "last_name": str(get_val("last_name") or "").strip(),
            "middle_name": str(get_val("middle_name") or "").strip(),
            "date_of_birth": self._parse_date(get_val("date_of_birth")),
            "gender": self._normalize_gender(get_val("gender")),
            "blood_group": str(get_val("blood_group") or "").strip(),
            "religion": str(get_val("religion") or "").strip(),
            "state_of_origin": str(get_val("state_of_origin") or "").strip(),
            "residential_address": str(get_val("residential_address") or "").strip(),
            "is_orphan": self._parse_bool(get_val("is_orphan")),
            "has_disability": self._parse_bool(get_val("has_disability")),
            "disability_details": str(get_val("disability_details") or "").strip(),
            "student_type": self._normalize_student_type(get_val("student_type")),
            "admission_date": self._parse_date(get_val("admission_date")),
            "academic_session": str(get_val("academic_session") or "2026").strip(),
            # Serializer field names, not *_id — DRF silently drops unknown keys,
            # which is why class and donor never used to come through.
            "student_class": self._parse_student_class(get_val("student_class")),
            "previous_school": str(get_val("previous_school") or "").strip(),
            "previous_class": str(get_val("previous_class") or "").strip(),
            "status": self._normalize_status(get_val("status")),
            "fee_status": self._normalize_fee_status(get_val("fee_status")),
            "donor": self._parse_donor(get_val("donor")),
            "donor_number": str(get_val("donor_number") or "").strip(),
        }

        parent_name = str(get_val("parent_name") or "").strip()
        parent_phone = str(get_val("parent_phone") or "").strip()
        relationship = str(get_val("relationship") or "").strip()
        if parent_name or parent_phone or relationship:
            if not parent_name or not parent_phone or not relationship:
                raise ValidationError("Parent data must include name, phone, and relationship.")
            payload["parent"] = {
                "full_name": parent_name,
                "phone": parent_phone,
                "relationship": relationship,
                "email": str(get_val("parent_email") or "").strip(),
                "occupation": str(get_val("occupation") or "").strip(),
                "office_address": str(get_val("office_address") or "").strip(),
                "home_address": str(get_val("home_address") or "").strip(),
            }
            second_name = str(get_val("guardian2_name") or "").strip()
            second_phone = str(get_val("guardian2_phone") or "").strip()
            second_relationship = str(get_val("guardian2_relationship") or "").strip()
            if second_name or second_phone or second_relationship:
                if not second_name or not second_phone or not second_relationship:
                    raise ValidationError(
                        "Second guardian data must include guardian2_name, guardian2_phone, and guardian2_relationship."
                    )
                payload["guardians"] = [
                    {
                        "full_name": second_name,
                        "phone": second_phone,
                        "relationship": second_relationship,
                        "email": str(get_val("guardian2_email") or "").strip(),
                        "occupation": str(get_val("guardian2_occupation") or "").strip(),
                        "is_primary": False,
                    }
                ]

            parent_user_id = get_val("parent_user_id")
            if str(parent_user_id).strip():
                try:
                    payload["parent_user_id"] = int(str(parent_user_id).strip())
                except ValueError:
                    raise ValidationError("parent_user_id must be an integer")

        if not payload["reg_no"]:
            raise ValidationError("reg_no is required")
        if not payload["first_name"]:
            raise ValidationError("first_name is required")
        if not payload["last_name"]:
            raise ValidationError("last_name is required")
        if not payload["date_of_birth"]:
            raise ValidationError("date_of_birth is required")
        if not payload["gender"]:
            raise ValidationError("gender is required")
        if not payload["admission_date"]:
            raise ValidationError("admission_date is required")

        return payload

    def _normalize_gender(self, value):
        if not value:
            return ""
        value_str = str(value).strip().lower()
        if value_str in ["male", "m"]:
            return "Male"
        if value_str in ["female", "f"]:
            return "Female"
        raise ValidationError(f"Gender must be Male or Female: {value}")

    def _normalize_student_type(self, value):
        if not value:
            return "Day"
        normalized = str(value).strip().lower()
        if normalized in ["day", "d"]:
            return "Day"
        if normalized in ["boarding", "b"]:
            return "Boarding"
        raise ValidationError(f"Student type must be Day or Boarding: {value}")

    def _normalize_status(self, value):
        if not value:
            return "active"
        normalized = str(value).strip().lower()
        if normalized in ["active"]:
            return "active"
        if normalized in ["suspended"]:
            return "suspended"
        if normalized in ["graduated"]:
            return "graduated"
        if normalized in ["withdrawn"]:
            return "withdrawn"
        raise ValidationError(f"Invalid status: {value}")

    def _normalize_fee_status(self, value):
        if not value:
            return "unpaid"
        normalized = str(value).strip().lower()
        if normalized in ["paid"]:
            return "paid"
        if normalized in ["partial"]:
            return "partial"
        if normalized in ["unpaid"]:
            return "unpaid"
        raise ValidationError(f"Invalid fee_status: {value}")

    @action(detail=False, methods=["post"], url_path="import-students")
    def import_students(self, request):
        """
        Bulk-create students from a .csv or .xlsx upload.

        Send ``dry_run=true`` to check the file and get a row-by-row preview
        without writing anything; send it again without the flag to save.
        Saving is all-or-nothing across the valid rows, so a failure part way
        through never leaves a half-imported batch behind.
        """
        dry_run = str(request.data.get("dry_run", "")).strip().lower() in ("1", "true", "yes", "on")

        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Please choose a CSV or Excel file to upload."}, status=status.HTTP_400_BAD_REQUEST)

        if not file.name.lower().endswith((".csv", ".xlsx")):
            return Response(
                {"detail": "Unsupported file type. Upload a .xlsx or .csv file — .xls is not supported, re-save it as .xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            rows = self._parse_import_rows(file)
        except ValidationError as exc:
            return Response({"detail": " ".join(exc.messages)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            return Response(
                {"detail": f"That file could not be read ({exc}). Make sure it is a valid .xlsx or .csv file and is not open in Excel."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not rows:
            return Response(
                {"detail": "No data rows found. The first row should hold the column headings, with one student per row underneath."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Validate every row before writing anything ───────────────────
        preview = []
        ready = []
        seen_reg_nos = {}

        for number, row in rows:
            entry = {"row": number, "reg_no": "", "name": "", "student_class": "", "valid": False, "errors": []}
            try:
                payload = self._build_import_payload(row)
            except ValidationError as exc:
                entry["reg_no"] = str(row.get("reg_no") or "").strip()
                entry["errors"] = list(exc.messages)
                preview.append(entry)
                continue
            except Exception as exc:
                entry["errors"] = [str(exc)]
                preview.append(entry)
                continue

            entry["reg_no"] = payload.get("reg_no", "")
            entry["name"] = f"{payload.get('first_name', '')} {payload.get('last_name', '')}".strip()
            entry["student_class"] = str(row.get("student_class") or "").strip()

            # A reg_no repeated inside the file would only blow up on the second
            # insert, so catch it here where it can be reported against the row.
            duplicate_of = seen_reg_nos.get(payload["reg_no"].lower())
            if duplicate_of:
                entry["errors"] = [f"reg_no '{payload['reg_no']}' is already used on row {duplicate_of} of this file."]
                preview.append(entry)
                continue

            serializer = StudentCreateSerializer(data=payload)
            if serializer.is_valid():
                seen_reg_nos[payload["reg_no"].lower()] = number
                entry["valid"] = True
                preview.append(entry)
                ready.append((number, payload))
            else:
                entry["errors"] = [
                    f"{field}: {'; '.join(str(m) for m in messages)}"
                    for field, messages in serializer.errors.items()
                ]
                preview.append(entry)

        valid_count = len(ready)
        invalid = [{"row": e["row"], "errors": e["errors"]} for e in preview if not e["valid"]]

        body = {
            "dry_run": dry_run,
            "total_rows": len(preview),
            "valid": valid_count,
            "invalid": len(invalid),
            "created": 0,
            "rows": preview,
            "errors": invalid,
        }

        if dry_run:
            return Response(body, status=status.HTTP_200_OK)

        if not valid_count:
            body["detail"] = "Nothing was saved — none of the rows are valid yet."
            return Response(body, status=status.HTTP_400_BAD_REQUEST)

        # ── Save: all valid rows together, or none at all ────────────────
        try:
            with transaction.atomic():
                for _, payload in ready:
                    serializer = StudentCreateSerializer(data=payload)
                    serializer.is_valid(raise_exception=True)
                    serializer.save()
        except Exception as exc:
            body["detail"] = f"Nothing was saved — the import was rolled back after an error: {exc}"
            return Response(body, status=status.HTTP_400_BAD_REQUEST)

        body["created"] = valid_count
        return Response(body, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["get"], url_path="import-columns")
    def import_columns(self, request):
        """The column specification the importer actually enforces."""
        from .import_spec import COLUMNS

        return Response({"columns": COLUMNS})

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """Download a ready-made .xlsx with the headings and a worked example."""
        from django.http import HttpResponse

        from .import_spec import build_template_workbook

        try:
            workbook = build_template_workbook()
        except ImportError:
            return Response(
                {"detail": "Excel support is not installed on the server (openpyxl)."},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="student-import-template.xlsx"'
        return response

    @action(detail=True, methods=["post"])
    def upload_document(self, request, pk=None):
        """Upload student document - Only admins."""
        if get_user_role(request.user) not in ['super_admin', 'admin']:
            return Response(
                {"error": "Only administrators can upload documents."},
                status=status.HTTP_403_FORBIDDEN
            )
        student = self.get_object()
        serializer = StudentDocumentSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        serializer.save(student=student)
        return Response(
            StudentDocumentSerializer(serializer.instance, context={"request": request}).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=["post"])
    def promote(self, request, pk=None):
        """
        Move student to a new class. Supports all movement types:
          - promotion:     move up (e.g. Form 1 → Form 2)
          - stream_change: same level different section (e.g. Form 1A → Form 1B)
          - repetition:    stay in same class (fail / attendance)
          - demotion:      move down (parent/school decision)
        Records the movement in AcademicHistory.
        Only admins can perform this action.
        """
        student = self.get_object()
        new_class_id = request.data.get("new_class")
        movement_type = request.data.get("movement_type", "promotion")
        reason = request.data.get("reason", "")
        academic_session = request.data.get("academic_session", "")
        term = request.data.get("term", "")

        VALID_MOVEMENTS = ["promotion", "stream_change", "repetition", "demotion"]
        if movement_type not in VALID_MOVEMENTS:
            return Response({"error": f"movement_type must be one of: {', '.join(VALID_MOVEMENTS)}"}, status=400)

        if not new_class_id:
            return Response({"error": "new_class is required"}, status=400)

        from academics.models import Class
        from .models import AcademicHistory
        from core.models import SchoolSettings

        try:
            new_class = Class.objects.get(pk=new_class_id)
        except Class.DoesNotExist:
            return Response({"error": "Class not found"}, status=404)

        # Resolve session/term if not provided
        if not academic_session or not term:
            cfg = SchoolSettings.objects.first()
            academic_session = academic_session or (cfg.academic_session if cfg else "2026")
            term = term or (cfg.current_term if cfg else "Term 3")

        old_class_name = student.class_name

        # Record academic history before moving
        try:
            from exams.models import ExamResult
            result = ExamResult.objects.filter(
                student=student,
                academic_session=academic_session,
                term=term,
            ).first()
            AcademicHistory.objects.update_or_create(
                student=student,
                term=term,
                academic_session=academic_session,
                defaults={
                    "position": result.position if result else 0,
                    "average": result.average if result else 0,
                    "grade": result.grade if result else "",
                    "division": result.division if result else "",
                    "previous_class": old_class_name,
                    "new_class": new_class.name,
                    "movement_type": movement_type,
                    "reason": reason,
                },
            )
        except Exception:
            pass  # History recording is best-effort; never block the move

        # Perform the move
        student.student_class = new_class
        if movement_type == "repetition":
            student.status = "active"
        elif movement_type == "demotion":
            student.status = "active"
        student.save()

        return Response({
            "detail": f"Student moved to {new_class.name} ({movement_type})",
            **StudentDetailSerializer(student).data,
        })


class AcademicHistoryViewSet(viewsets.ModelViewSet):
    """
    Academic history per student - Admins manage, teachers/parents can read their own.
    """
    queryset = AcademicHistory.objects.select_related("student").all()
    serializer_class = AcademicHistorySerializer
    permission_classes = [IsTeacherOrAdminOrParentReadOnly]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["student", "term", "academic_session"]
    ordering_fields = ["academic_session", "term"]
    ordering = ["-academic_session", "term"]

    def get_queryset(self):
        user = self.request.user
        role = get_user_role(user)
        if role in ['super_admin', 'admin']:
            return AcademicHistory.objects.select_related("student").all()
        elif role == 'teacher':
            all_ids = get_teacher_class_ids(get_teacher_for(user))
            return AcademicHistory.objects.filter(
                student__student_class_id__in=all_ids
            ).select_related("student")
        elif role == 'parent':
            return AcademicHistory.objects.filter(
                student__guardians__user=user
            ).select_related("student").distinct()
        return AcademicHistory.objects.none()

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsSuperAdminOrAdmin()]
        return [IsTeacherOrAdminOrParentReadOnly()]
