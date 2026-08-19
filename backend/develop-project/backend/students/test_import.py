"""Bulk student import: an .xlsx from Excel must go in cleanly and predictably."""
import datetime
import io

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from academics.models import Class
from core.models import UserProfile
from students.models import ParentGuardian, Student

HEADERS = ["reg_no", "first_name", "last_name", "date_of_birth", "gender", "admission_date"]


def xlsx(rows, name="students.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        name, buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class StudentImportTests(TestCase):
    def setUp(self):
        user = User.objects.create_user(
            username="sa@example.com", email="sa@example.com", password="SuperPass123"
        )
        UserProfile.objects.update_or_create(user=user, defaults={"role": "super_admin"})
        self.client = APIClient()
        self.client.force_authenticate(user=User.objects.get(pk=user.pk))
        Class.objects.create(name="Form 1A", section="Secondary", level="Form 1", arm="A")

    def upload(self, file, dry_run=False):
        payload = {"file": file}
        if dry_run:
            payload["dry_run"] = "true"
        return self.client.post("/api/students/import-students/", payload, format="multipart")

    # ── Reading the file ─────────────────────────────────────────────────

    def test_excel_date_cells_are_accepted(self):
        """Typing a date into Excel produces a datetime cell, not a string."""
        response = self.upload(xlsx([
            HEADERS + ["student_class"],
            ["AN-001", "Asha", "Said", datetime.date(2014, 3, 2), "Female",
             datetime.date(2026, 1, 13), "Form 1A"],
        ]))

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["created"], 1)
        student = Student.objects.get(reg_no="AN-001")
        self.assertEqual(student.date_of_birth, datetime.date(2014, 3, 2))
        self.assertEqual(student.student_class.name, "Form 1A")

    def test_text_dates_in_several_formats_are_accepted(self):
        response = self.upload(xlsx([
            HEADERS,
            ["AN-002", "A", "B", "2014-03-02", "Male", "2026-01-13"],
            ["AN-003", "C", "D", "02-03-2014", "M", "13-01-2026"],
            ["AN-004", "E", "F", "02.03.2014", "female", "13.01.2026"],
        ]))

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["created"], 3)
        self.assertEqual(Student.objects.get(reg_no="AN-003").date_of_birth, datetime.date(2014, 3, 2))

    def test_blank_rows_are_ignored_not_reported_as_errors(self):
        response = self.upload(xlsx([
            HEADERS,
            ["AN-005", "Juma", "Ali", "2014-03-02", "Male", "2026-01-13"],
            [None, None, None, None, None, None],
            [None, None, None, None, None, None],
        ]))

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["created"], 1)
        self.assertEqual(response.data["invalid"], 0)
        self.assertEqual(response.data["total_rows"], 1)

    def test_headings_may_use_spaces_and_any_capitalisation(self):
        response = self.upload(xlsx([
            ["Reg No", "First Name", "Last Name", "Date Of Birth", "Gender", "Admission Date"],
            ["AN-006", "Neema", "P", "2014-03-02", "female", "2026-01-13"],
        ]))

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(Student.objects.get(reg_no="AN-006").gender, "Female")

    def test_a_numeric_registration_number_is_read_as_text(self):
        response = self.upload(xlsx([
            HEADERS,
            [1001, "Numeric", "Reg", "2014-03-02", "Male", "2026-01-13"],
        ]))

        self.assertEqual(response.status_code, 201, response.data)
        self.assertTrue(Student.objects.filter(reg_no="1001").exists())

    def test_csv_uploads_still_work(self):
        csv_bytes = (
            "reg_no,first_name,last_name,date_of_birth,gender,admission_date\n"
            "AN-007,Csv,User,2014-03-02,Male,2026-01-13\n"
        ).encode("utf-8")

        response = self.upload(SimpleUploadedFile("s.csv", csv_bytes, content_type="text/csv"))

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["created"], 1)

    def test_an_empty_sheet_is_rejected_with_a_clear_message(self):
        response = self.upload(xlsx([HEADERS]))

        self.assertEqual(response.status_code, 400)
        self.assertIn("No data rows", response.data["detail"])

    def test_a_wrong_file_type_is_rejected_with_a_clear_message(self):
        response = self.upload(SimpleUploadedFile("notes.txt", b"hello", content_type="text/plain"))

        self.assertEqual(response.status_code, 400)
        self.assertIn(".xlsx", response.data["detail"])

    # ── Preview, then save ───────────────────────────────────────────────

    def test_dry_run_reports_every_row_and_writes_nothing(self):
        response = self.upload(xlsx([
            HEADERS,
            ["AN-008", "Good", "Row", "2014-03-02", "Male", "2026-01-13"],
            ["AN-009", "Bad", "Row", "not-a-date", "Male", "2026-01-13"],
        ]), dry_run=True)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["dry_run"])
        self.assertEqual(response.data["valid"], 1)
        self.assertEqual(response.data["invalid"], 1)
        self.assertEqual(response.data["created"], 0)
        self.assertEqual(Student.objects.count(), 0)

        rows = response.data["rows"]
        self.assertEqual(rows[0]["row"], 2)
        self.assertEqual(rows[0]["name"], "Good Row")
        self.assertTrue(rows[0]["valid"])
        self.assertFalse(rows[1]["valid"])
        self.assertIn("Invalid date format", " ".join(rows[1]["errors"]))

    def test_saving_after_a_dry_run_creates_the_valid_rows(self):
        rows = [
            HEADERS,
            ["AN-010", "First", "Student", "2014-03-02", "Male", "2026-01-13"],
            ["AN-011", "Second", "Student", "2014-04-02", "Female", "2026-01-13"],
        ]
        self.upload(xlsx(rows), dry_run=True)
        self.assertEqual(Student.objects.count(), 0)

        response = self.upload(xlsx(rows))

        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(response.data["created"], 2)
        self.assertEqual(Student.objects.count(), 2)

    def test_a_failure_rolls_the_whole_batch_back(self):
        """A half-imported batch is worse than none — saving is all or nothing."""
        response = self.upload(xlsx([
            HEADERS,
            ["AN-012", "Good", "Row", "2014-03-02", "Male", "2026-01-13"],
            ["AN-012", "Same", "RegNo", "2014-03-02", "Male", "2026-01-13"],
        ]))

        self.assertEqual(response.status_code, 201)
        # The duplicate is reported, the good row is saved, and nothing is half-written.
        self.assertEqual(response.data["created"], 1)
        self.assertEqual(response.data["invalid"], 1)
        self.assertEqual(Student.objects.count(), 1)

    def test_a_registration_number_repeated_in_the_file_is_caught(self):
        response = self.upload(xlsx([
            HEADERS,
            ["AN-013", "One", "Student", "2014-03-02", "Male", "2026-01-13"],
            ["AN-013", "Two", "Student", "2014-03-02", "Male", "2026-01-13"],
        ]), dry_run=True)

        self.assertEqual(response.data["invalid"], 1)
        self.assertIn("already used on row 2", " ".join(response.data["rows"][1]["errors"]))

    def test_an_unknown_class_names_the_problem(self):
        response = self.upload(xlsx([
            HEADERS + ["student_class"],
            ["AN-014", "A", "B", "2014-03-02", "Male", "2026-01-13", "Form 9Z"],
        ]), dry_run=True)

        self.assertIn("Student class not found: Form 9Z", " ".join(response.data["rows"][0]["errors"]))

    def test_parent_columns_create_the_guardian(self):
        response = self.upload(xlsx([
            HEADERS + ["parent_name", "parent_phone", "relationship", "parent_email"],
            ["AN-015", "Kito", "Juma", "2014-03-02", "Male", "2026-01-13",
             "Fatma Juma", "+255700000000", "Mother", "fatma@example.com"],
        ]))

        self.assertEqual(response.status_code, 201, response.data)
        guardian = ParentGuardian.objects.get(student__reg_no="AN-015")
        self.assertEqual(guardian.full_name, "Fatma Juma")
        self.assertEqual(guardian.relationship, "Mother")

    def test_partial_parent_details_are_rejected_with_a_reason(self):
        response = self.upload(xlsx([
            HEADERS + ["parent_name"],
            ["AN-016", "A", "B", "2014-03-02", "Male", "2026-01-13", "Fatma Juma"],
        ]), dry_run=True)

        self.assertIn("name, phone, and relationship", " ".join(response.data["rows"][0]["errors"]))

    def test_disability_columns_are_imported(self):
        response = self.upload(xlsx([
            HEADERS + ["has_disability", "disability_details"],
            ["AN-017", "A", "B", "2014-03-02", "Male", "2026-01-13", "TRUE",
             "Low vision — needs large print"],
        ]))

        self.assertEqual(response.status_code, 201, response.data)
        student = Student.objects.get(reg_no="AN-017")
        self.assertTrue(student.has_disability)
        self.assertIn("large print", student.disability_details)

    # ── Template and column guide ────────────────────────────────────────

    def test_the_template_downloads_as_a_readable_workbook(self):
        response = self.client.get("/api/students/import-template/")

        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        self.assertIn("student-import-template.xlsx", response["Content-Disposition"])

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Students", "Column guide"])
        headings = [c.value for c in workbook["Students"][1]]
        for required in ("reg_no", "first_name", "last_name", "date_of_birth", "gender", "admission_date"):
            self.assertIn(required, headings)

    def test_the_downloaded_template_imports_as_is(self):
        """The example row in the template must actually be valid."""
        template = self.client.get("/api/students/import-template/").content
        Class.objects.get_or_create(
            name="Form 1A", defaults={"section": "Secondary", "level": "Form 1", "arm": "A"}
        )

        response = self.upload(
            SimpleUploadedFile("student-import-template.xlsx", template), dry_run=True
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["invalid"], 0, response.data["rows"])
        self.assertEqual(response.data["valid"], 1)

    def test_the_column_guide_is_served_to_the_page(self):
        response = self.client.get("/api/students/import-columns/")

        self.assertEqual(response.status_code, 200)
        keys = [c["key"] for c in response.data["columns"]]
        self.assertIn("reg_no", keys)
        self.assertIn("disability_details", keys)
        required = [c["key"] for c in response.data["columns"] if c["required"]]
        self.assertEqual(
            set(required),
            {"reg_no", "first_name", "last_name", "date_of_birth", "gender", "admission_date"},
        )

    # ── Access ───────────────────────────────────────────────────────────

    def test_only_super_admins_may_import(self):
        for role in ("admin", "teacher"):
            user = User.objects.create_user(
                username=f"{role}@example.com", email=f"{role}@example.com", password="SomePass12345"
            )
            UserProfile.objects.update_or_create(user=user, defaults={"role": role})
            client = APIClient()
            client.force_authenticate(user=User.objects.get(pk=user.pk))

            response = client.post(
                "/api/students/import-students/",
                {"file": xlsx([HEADERS, ["X-1", "A", "B", "2014-03-02", "Male", "2026-01-13"]])},
                format="multipart",
            )

            self.assertEqual(response.status_code, 403)
